import os
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from docx2pdf import convert

workbook = load_workbook(filename="Module overzicht Technische Informatica.xlsx")
sheets   = workbook.sheetnames
sheet    = workbook["Modulewijzer"] #workbook.active

header  = None
records = []

for row in sheet.iter_rows(min_row=1, values_only=True):
    if ( header == None ):
        header = row
        print(header)

    else:
        record = {}
        for i in range(0, len(header)):
            record[header[i]] = str( row[i] ).strip()
        records.append( record )

doc = DocxTemplate("Template Modulewijzer.docx")

for record in records:
    doc.render(record)
    doc.save("Modulewijzer " + record["Module"] + ".docx")

#convert("generated.docx", os.getcwd() + "/generated.pdf")
