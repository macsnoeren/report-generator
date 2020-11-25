import os
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from docx2pdf import convert

workbook = load_workbook(filename="Beoordeling OGP0 Totaal v1.xlsx")
sheets   = workbook.sheetnames
sheet    = workbook["Feedback"] #workbook.active

header  = None
records = []

for row in sheet.iter_rows(min_row=4, values_only=True):
    if ( header == None ):
        header = row
        print(header)

    else:
        feedbacks = []
        for i in range(3, 42, 2):
                feedback = { "Opgave": str( header[i] ), "Punten": str(row[i]), "Max": str(header[i+1]), "Feedback": str(row[i+1]) }
                feedbacks.append(feedback)
        
        record = { "Student": str( row[0] ), "Punten": str(row[44]), "Cijfer": str(row[45]), "Feedbacks": feedbacks }    
        records.append( record )

for record in records:
    doc = DocxTemplate("Template Student Feedback.docx")
    doc.render(record)
    print("---\n")
    print(record)
    print("\n---\n")
    doc.save("output\Feedback " + record["Student"] + ".docx")

#convert("generated.docx", os.getcwd() + "/generated.pdf")
