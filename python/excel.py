# https://realpython.com/openpyxl-excel-spreadsheets-python/

from openpyxl import load_workbook

workbook = load_workbook(filename="Ecovat SCADA Test Book.xlsx")
sheets   = workbook.sheetnames
sheet    = workbook["Tests cases"] #workbook.active

header = None

# Using the values_only because you want to return the cells' values
records = []
for row in sheet.iter_rows(min_row=1, values_only=True):
    if ( header == None ):
        header = row

    else:
        record = {}
        for i in range(1, len(header)):
            record[header[i]] = row[i]
        records.append( record )

print(records)